import os
import pytest
import unittest

import re

from bcompiler.compile import parse_source_cells, get_current_quarter
from bcompiler.datamap import Datamap, DatamapLine

from bcompiler.utils import VALIDATION_REFERENCES, SHEETS
from bcompiler.utils import index_returns_directory

from openpyxl import Workbook


@pytest.mark.skip(reason='Need to deal with BOM')
class TestCompilationFromReturns(unittest.TestCase):
    def setUp(self):
        self.cell_regex = re.compile('[A-Z]+[0-9]+')
        self.dropdown_regex = re.compile('^\D*$')
        self.docs = os.path.join(os.path.expanduser('~'), 'Documents')
        bcomp_working_d = 'bcompiler'
        self.path = os.path.join(self.docs, bcomp_working_d)
        self.source_path = os.path.join(self.path, 'source')
        self.output_path = os.path.join(self.path, 'output')
        self.returns_path = os.path.join(self.source_path, 'returns')
        self.source_file_name = 'Crossrail Programme_Q3_Return.xlsx'
        self.source_excel = os.path.join(
            self.returns_path, self.source_file_name)
        self.datamap_returns_to_master = os.path.join(
            self.source_path, 'datamap-returns-to-master')
        self.datamap_master_to_returns = os.path.join(
            self.source_path, 'datamap-master-to-returns')
        self.dm = Datamap(
            datamap_type='returns-to-master',
            source_file=self.datamap_returns_to_master)
        self.example_return = os.path.join(
            self.source_path, "returns/Crossrail Programme_Q3_Return.xlsx")

    @pytest.mark.skip(reason='Need to deal with BOM')
    def test_parse_source_excel_file(self):
        parsed_data = parse_source_cells(
            self.source_excel,
            self.datamap_returns_to_master)
        self.assertEqual('Project/Programme Name', parsed_data[0]['gmpp_key'])
        self.assertEqual('DVSA IT Sourcing', parsed_data[0]['gmpp_key_value'])

    @pytest.mark.skip(reason='Need to deal with BOM')
    def test_get_quarter(self):
        self.assertEqual(get_current_quarter(
            self.source_file_name), 'Q1 Apr - Jun')

    @pytest.mark.skip(reason='Need to deal with BOM')
    def test_dropdown_not_passing_to_master_bug(self):
        return_f = self.example_return
        data = parse_source_cells(return_f, self.datamap_master_to_returns)
        example_validated_cell = "IO1 - Monetised?"
        matches = [x for x in data if x['gmpp_key'] == example_validated_cell]
        self.assertEqual(matches[0]['gmpp_key'], example_validated_cell)

    @pytest.mark.skip(reason='Need to deal with BOM')
    def test_parse_returned_form(self):
        return_f = self.example_return
        data = parse_source_cells(return_f, self.datamap_master_to_returns)
        self.assertEqual(data[0]['gmpp_key'], 'Project/Programme Name')


@pytest.mark.skip(reason="Fragile test")
class TestDatamapFunctionality(unittest.TestCase):
    def setUp(self):
        self.cell_regex = re.compile('[A-Z]+[0-9]+')
        self.dropdown_regex = re.compile('^\D*$')
        self.docs = os.path.join(os.path.expanduser('~'), 'Documents')
        bcomp_working_d = 'bcompiler'
        self.path = os.path.join(self.docs, bcomp_working_d)
        self.source_path = os.path.join(self.path, 'source')
        self.output_path = os.path.join(self.path, 'output')
        self.datamap_master_to_returns = os.path.join(
            self.source_path, 'datamap-master-to-returns')
        self.datamap_returns_to_master = os.path.join(
            self.source_path, 'datamap-returns-to-master')
        self.master = os.path.join(self.source_path, 'master.csv')
        self.transposed_master = os.path.join(
            self.source_path, 'master_transposed.csv')
        self.dm = Datamap(
            datamap_type='returns-to-master',
            source_file=self.datamap_returns_to_master)

    @pytest.mark.skip(reason="Fragile test")
    def test_verified_lines(self):
        # these are DatamapLine objects that have 4 attributes, the last of
        # which is verification dropdown text
        # the last element in each should be a dropdown text string
        for item in self.dm._dml_cname_sheet_cref_ddown:
            self.assertTrue(self.dropdown_regex, item.dropdown_txt)

    @pytest.mark.skip(reason="Fragile test")
    def test_verified_lines_for_dropdown_text(self):
        # we're expecting the dropdown_txt attr in the DatamapLine object
        # to be what we expect
        for item in self.dm._dml_cname_sheet_cref_ddown:
            self.assertTrue(item.dropdown_txt in VALIDATION_REFERENCES.keys())

    @pytest.mark.skip(reason="Fragile test")
    def test_non_verified_lines(self):
        # these are DatamapLine objects that have 3 attributes, the
        # last of which is a regex
        for item in self.dm._dml_cname_sheet_cref:
            self.assertTrue(self.cell_regex, item.cellref)

    @pytest.mark.skip(reason="We don't have a GMPP info sheet yet")
    def test_cells_that_will_not_migrate(self):
        # these are DatamapLine objects that have 2 attributes,
        # the last of which is a sheet ref
        for item in self.dm._dml_cname_sheet:
            self.assertTrue(item.sheet in SHEETS)

    @pytest.mark.skip(reason="Fragile test")
    def test_single_item_lines(self):
        # DatamapLines that have a single attribute
        for item in self.dm._dml_cname:
            # TODO this is fragile - shouldn't be counting lines in this test
            self.assertEqual(self.dm.count_dml_cellname_only, 18)

    @pytest.mark.skip(reason="Fragile test")
    def test_datamap_is_cleaned_attr(self):
        self.assertTrue(self.dm.is_cleaned)

    @pytest.mark.skip(reason="Fragile test")
    def test_pretty_dataline_print(self):
        dml = DatamapLine()
        dml.cellname = 'Test cellname'
        dml.sheet = 'Summary'
        dml.cellref = 'C12'
        dml.dropdown_txt = 'Finance Figures'
        self.assertEqual(
            dml.pretty_print(), ("Name: Test cellname | Sheet: Summary | "
                                 "Cellref: C12 | Dropdown: Finance Figures"))

    @pytest.mark.skip(reason="Fragile test")
    def test_index_returns_directory(self):
        assert index_returns_directory() == []


# LATER PYTEST-ONLY TESTS FOR COMPILATION PROCESS, USING FIXTURES

@pytest.fixture
def bicc_return():
    wb = Workbook()
    wb.create_sheet('Summary')
    wb.create_sheet('Approval & Project milestones')
    wb.create_sheet('Finance & Benefits')
    wb.create_sheet('Resources')
    wb.create_sheet('Assurance planning')
    wb.create_sheet('GMPP info')
    ws_summary = wb['Summary']
    # enter some values in the right slots
    ws_summary['B5'].value = 'Cookfield Rebuild'
    ws_summary['B8'].value = 'Roads, Monitoring and Horse'
    ws_summary['C12'].value = 'Keith McGilvrey'
    ws_summary['C13'].value = '0208 944 3554'
    ws_summary['C14'].value = 'cleft.palet@screen.co.ok'
    ws_summary['C21'].value = 'Cohort 13'
    ws_summary['B26'].value = 'We think this is a really good idea, ok?'

    wb.save('/tmp/test-bicc-return.xlsx')
    yield '/tmp/test-bicc-return.xlsx'
    os.unlink('/tmp/test-bicc-return.xlsx')


@pytest.fixture
def old_master():
    wb = Workbook()
    wb.create_sheet('Constructed BICC Data Master')
    ws = wb['Constructed BICC Data Master']
    ws['A1'].value = 'Project/Programme Name'
    ws['A2'].value = 'SRO Sign-Off'
    # TODO carry on with this
